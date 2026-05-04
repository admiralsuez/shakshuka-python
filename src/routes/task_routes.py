"""
Task Routes - All task management endpoints

This module handles:
- Task CRUD operations (GET, POST, PUT, DELETE)
- Task actions (complete, strike, undo-strike, schedule, unschedule)
- Task imports from CSV/TXT files
- Daily strike resets
"""

from flask import Blueprint, request, jsonify, current_app, send_file
import logging
import uuid
import io
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
import re

# Import app context and utilities (will be injected)
from src.constants import DEFAULT_USER_ID, TaskStatus
from src.services.importer import parse_csv_tasks, parse_txt_tasks
from src.exceptions import DatabaseError
from src.routes.api_utils import register_api_error_handlers
from src.services import scheduler as scheduler_service

# Import decorators and validators
from src.routes.route_decorators import (
    require_data_manager,
    require_json_body,
    require_file_upload,
    validate_input,
    rate_limit,
    handle_database_error
)
from src.routes.input_validators import (
    validate_schedule_input,
    validate_task_title,
    validate_priority,
    validate_date_yyyy_mm_dd,
    validate_description,
    validate_project_name,
    validate_owner_name,
    validate_strike_report,
    validate_bulk_operation_count
)

logger = logging.getLogger(__name__)

# Blueprint definition
task_bp = Blueprint('tasks', __name__, url_prefix='/api/tasks')

register_api_error_handlers(task_bp)

# These will be injected at runtime
_app_context = None
_get_user_id_func = None
_ensure_data_manager_func = None
_sanitize_input_func = None
_validate_task_data_func = None
_rate_limit_decorator = None


def init_task_routes(app_context, get_user_id_func, ensure_data_manager_func, 
                     sanitize_input_func, validate_task_data_func, rate_limit_decorator):
    """Initialize task routes with dependency injection"""
    global _app_context, _get_user_id_func, _ensure_data_manager_func
    global _sanitize_input_func, _validate_task_data_func, _rate_limit_decorator
    
    _app_context = app_context
    _get_user_id_func = get_user_id_func
    _ensure_data_manager_func = ensure_data_manager_func
    _sanitize_input_func = sanitize_input_func
    _validate_task_data_func = validate_task_data_func
    _rate_limit_decorator = rate_limit_decorator


def _get_user_id():
    """Get current user ID"""
    if _get_user_id_func:
        return _get_user_id_func()
    return DEFAULT_USER_ID


def _get_data_manager():
    """Get data manager instance"""
    if _ensure_data_manager_func and not _ensure_data_manager_func():
        return None
    ctx = _app_context
    if ctx is None:
        try:
            ctx = current_app.extensions.get('app_context')
        except Exception:  # noqa: broad-except
            ctx = None
    return ctx.data_manager if ctx else None


_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


# Validator functions for decorators
def validate_schedule(data):
    """Validate schedule input"""
    hour = data.get('hour')
    minute = data.get('minute', 0)
    duration = data.get('duration', 30)
    return validate_schedule_input(hour, minute, duration)


def validate_strike(data):
    """Validate strike input"""
    strike_type = data.get('type')
    report = data.get('report', '')
    
    if not strike_type or strike_type not in ['today', 'forever']:
        return False, "Invalid strike type"
    
    if not isinstance(report, str):
        return False, "Report must be a string"
    
    if len(report) > 2000:
        return False, "Report too long (max 2000 characters)"
    
    return True, ""


def validate_task_creation(data):
    """Validate task creation input"""
    title = data.get('title', '')
    valid, error = validate_task_title(title)
    if not valid:
        return False, error
    
    priority = data.get('priority', 'medium')
    if priority:
        valid, error = validate_priority(priority)
        if not valid:
            return False, error
    
    project = data.get('project', '')
    if project:
        valid, error = validate_project_name(project)
        if not valid:
            return False, error
    
    owner = data.get('owner', '')
    if owner:
        valid, error = validate_owner_name(owner)
        if not valid:
            return False, error
    
    description = data.get('description', '')
    if description:
        valid, error = validate_description(description)
        if not valid:
            return False, error
    
    return True, ""


def _validate_date_yyyy_mm_dd(value: str) -> bool:
    if not isinstance(value, str):
        return False
    if not _DATE_RE.match(value):
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def _parse_schedule_payload(payload: Dict) -> Tuple[bool, Dict, str]:
    if not isinstance(payload, dict):
        return False, {}, "Request must contain JSON object"

    hour_input = payload.get('hour')
    minute_input = payload.get('minute', 0)
    duration_input = payload.get('duration', 30)
    date_input = payload.get('date')

    if hour_input is None:
        return False, {}, "Hour is required"

    # Parse hour/minute
    hour: int
    minute: int

    if isinstance(hour_input, str) and ':' in hour_input:
        parts = hour_input.split(':', 1)
        try:
            hour = int(parts[0])
            minute = int(parts[1])
        except Exception:  # noqa: broad-except
            return False, {}, "Hour must be an integer hour or 'HH:MM'"
    else:
        try:
            hour = int(hour_input)
        except Exception:  # noqa: broad-except
            return False, {}, "Hour must be an integer"
        try:
            minute = int(minute_input)
        except Exception:  # noqa: broad-except
            return False, {}, "Minute must be an integer"

    if hour < 0 or hour > 23:
        return False, {}, "Hour must be between 0 and 23"
    if minute < 0 or minute > 59:
        return False, {}, "Minute must be between 0 and 59"

    # Duration
    try:
        duration = int(duration_input)
    except Exception:  # noqa: broad-except
        return False, {}, "Duration must be an integer"
    if duration < 5 or duration > 480:
        return False, {}, "Duration must be between 5 and 480 minutes"

    # Date (default to today in local time)
    if date_input is None:
        date_value = datetime.now().strftime('%Y-%m-%d')
    else:
        date_value = date_input

    if not _validate_date_yyyy_mm_dd(date_value):
        return False, {}, "Date must be in YYYY-MM-DD format"

    return True, {
        'hour': hour,
        'minute': minute,
        'duration': duration,
        'date': date_value,
    }, ""


@task_bp.route('', methods=['GET'])
@require_data_manager
@handle_database_error
def get_tasks(user_id, data_manager):
    """Get all tasks for the current user"""
    logger.info(f"API get_tasks called with user_id: {user_id}")
    
    tasks = data_manager.load_tasks(user_id)
    logger.info(f"Loaded {len(tasks)} tasks for user {user_id}")
    return jsonify(tasks)


@task_bp.route('/import', methods=['POST'])
def import_tasks():
    """Import tasks from CSV or TXT file with batch operations and rate limiting"""
    # Rate limiting: max 10 imports per hour per user
    user_id = _get_user_id()
    if not hasattr(import_tasks, '_import_times'):
        import_tasks._import_times = {}
    
    from datetime import datetime, timedelta
    now = datetime.now()
    key = user_id
    
    # Clean old requests (older than 1 hour)
    if key in import_tasks._import_times:
        import_tasks._import_times[key] = [
            t for t in import_tasks._import_times[key]
            if (now - t).total_seconds() < 3600
        ]
    else:
        import_tasks._import_times[key] = []
    
    # Check rate limit
    if len(import_tasks._import_times[key]) >= 10:
        return jsonify({
            'error': 'Rate limit exceeded',
            'message': 'Maximum 10 imports per hour',
            'retry_after': 3600
        }), 429
    
    # Record this import
    import_tasks._import_times[key].append(now)
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '' or not file:
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        # Read file content
        file_content = file.read().decode('utf-8')
        file_extension = file.filename.lower().split('.')[-1]
        
        imported_tasks = []
        errors = []
        
        if file_extension == 'csv':
            imported_tasks, errors = parse_csv_tasks(file_content, _sanitize_input_func)
        elif file_extension == 'txt':
            imported_tasks, errors = parse_txt_tasks(file_content, _sanitize_input_func)
        else:
            return jsonify({'error': 'Unsupported file format. Please use CSV or TXT.'}), 400
        
        if not imported_tasks:
            return jsonify({'error': 'No valid tasks found in file', 'details': errors}), 400
        
        # Limit to 1000 tasks per import
        if len(imported_tasks) > 1000:
            return jsonify({
                'error': 'Too many tasks',
                'message': 'Maximum 1000 tasks per import',
                'count': len(imported_tasks)
            }), 400
        
        # Get user and data manager
        user_id = _get_user_id()
        data_manager = _get_data_manager()
        if not data_manager:
            return jsonify({'error': 'Data manager not available'}), 500
        
        # Prepare tasks with IDs and timestamps
        tasks_to_import = []
        for task in imported_tasks:
            task['id'] = str(uuid.uuid4())
            task['created_at'] = datetime.now().isoformat()
            task['completed'] = False
            task['strike_count'] = 0
            task['struck_today'] = False
            tasks_to_import.append(task)
        
        # Use batch create instead of load-all/save-all (100x faster!)
        try:
            success = data_manager.bulk_create_tasks(user_id, tasks_to_import)
            if success:
                return jsonify({
                    'success': True,
                    'message': f'Successfully imported {len(tasks_to_import)} tasks',
                    'imported_count': len(tasks_to_import),
                    'errors': errors
                })
            else:
                return jsonify({'error': 'Failed to save imported tasks'}), 500
        except DatabaseError:
            logger.exception("Database error during bulk import for user %s", user_id)
            return jsonify({'error': 'Database error during import'}), 503
            
    except Exception:  # noqa: broad-except
        logger.exception("Import failed")
        return jsonify({'error': 'Import failed'}), 500


@task_bp.route('', methods=['POST'])
@require_data_manager
@require_json_body
@validate_input(validate_task_creation)
@handle_database_error
def create_task(user_id, data_manager):
    """Create a new task with comprehensive validation and error handling.

    When a task with the same (title, project) already exists for the user and
    is still active, the data manager will refuse creation and we return a
    409 Conflict so the UI can offer "Add again" / "Rename & add" options.
    """
    logger.info(f"API create_task called with user_id: {user_id}")
    
    task_data = request.json
    
    # Sanitize input data
    if _sanitize_input_func:
        task_data = _sanitize_input_func(task_data)
    
    # Create task using data manager
    created_task = data_manager.create_task_for_user(user_id, task_data)
    
    if created_task:
        logger.info(f"Successfully created task {created_task['id']} for user {user_id}")
        return jsonify(created_task), 201
    else:
        # Treat a failed creation here as a likely duplicate conflict so
        # the frontend can offer the user an explicit choice.
        logger.warning(f"Duplicate or failed task creation for user {user_id}; returning 409")
        return jsonify({'error': 'A similar task already exists'}), 409


@task_bp.route('/<task_id>', methods=['PUT'])
@require_data_manager
@require_json_body
@validate_input(validate_task_creation)
@handle_database_error
def update_task(task_id, user_id, data_manager):
    """Update an existing task with comprehensive validation and error handling"""
    logger.info(f"API update_task called for task {task_id} with user_id: {user_id}")
    
    task_data = request.json
    
    # Sanitize input data
    if _sanitize_input_func:
        task_data = _sanitize_input_func(task_data)
    
    # Update task using data manager
    success = data_manager.update_task_for_user(user_id, task_id, task_data)
    
    if success:
        # Track edit in analytics
        try:
            from src.analytics_manager import increment_analytics_counter
            increment_analytics_counter('tasks_edited')
        except Exception:  # noqa: broad-except
            logger.exception("Failed to increment analytics counter (tasks_edited)")
        # Return updated task directly from database
        updated_task = data_manager.get_task_by_id(user_id, task_id)
        if updated_task:
            logger.info(f"Successfully updated task {task_id} for user {user_id}")
            return jsonify(updated_task)
        
        logger.error(f"Task {task_id} not found after update for user {user_id}")
        return jsonify({'error': 'Task not found after update'}), 500
    else:
        logger.error(f"Failed to update task {task_id} for user {user_id}")
        return jsonify({'error': 'Failed to update task'}), 500


@task_bp.route('/<task_id>', methods=['DELETE'])
@require_data_manager
@handle_database_error
def delete_task(task_id, user_id, data_manager):
    """Delete a task with comprehensive error handling"""
    logger.info(f"API delete_task called for task {task_id} with user_id: {user_id}")
    
    if not task_id or not isinstance(task_id, str):
        return jsonify({'error': 'Invalid task ID'}), 400
    
    # Get task before deletion for response
    task_to_delete = data_manager.get_task_by_id(user_id, task_id)
    
    if not task_to_delete:
        logger.warning(f"Task {task_id} not found for user {user_id}")
        return jsonify({'error': 'Task not found'}), 404

    delete_source = request.headers.get('X-Delete-Source', '')
    if str(delete_source).lower() == 'tasklist':
        try:
            data_manager.save_deleted_task_snapshot(user_id, task_to_delete)
        except Exception:  # noqa: broad-except
            logger.exception("Failed to snapshot deleted task (tasklist source)")

    # Delete task using data manager
    success = data_manager.delete_task_for_user(user_id, task_id)
    
    if success:
        # Track deletion in analytics
        try:
            from src.analytics_manager import increment_analytics_counter
            if str(delete_source).lower() == 'tasklist':
                increment_analytics_counter('tasks_deleted')
        except Exception:  # noqa: broad-except
            logger.exception("Failed to increment analytics counter (tasks_deleted)")
        logger.info(f"Successfully deleted task {task_id} for user {user_id}")
        return jsonify(task_to_delete)
    else:
        logger.error(f"Failed to delete task {task_id} for user {user_id}")
        return jsonify({'error': 'Failed to delete task'}), 500


@task_bp.route('/<task_id>/undo-delete', methods=['POST'])
def undo_delete_task(task_id):
    user_id = _get_user_id()
    data_manager = _get_data_manager()
    if not data_manager:
        return jsonify({'error': 'Data manager not available'}), 500

    try:
        restored = data_manager.restore_deleted_task_snapshot(user_id, task_id)
        if not restored:
            return jsonify({'error': 'Nothing to restore'}), 404

        try:
            from src.analytics_manager import decrement_analytics_counter
            delete_source = request.headers.get('X-Delete-Source', '')
            if str(delete_source).lower() == 'tasklist':
                decrement_analytics_counter('tasks_deleted', 1)
        except Exception:  # noqa: broad-except - API route error handler must catch all exceptions
            logger.exception("Failed to decrement analytics counter (tasks_deleted)")

        return jsonify(restored), 200
    except Exception:  # noqa: broad-except
        logger.exception("Unexpected error in undo_delete_task for user %s, task %s", user_id, task_id)
        return jsonify({'error': 'Internal server error'}), 500


@task_bp.route('/<task_id>/strike-reports', methods=['GET'])
def get_strike_today_report_history(task_id):
    user_id = _get_user_id()
    data_manager = _get_data_manager()
    if not data_manager:
        return jsonify({'success': False, 'error': 'Data manager not available'}), 500

    limit_raw = request.args.get('limit', '200')
    offset_raw = request.args.get('offset', '0')
    try:
        limit = int(limit_raw)
    except (TypeError, ValueError):
        limit = 200
    try:
        offset = int(offset_raw)
    except (TypeError, ValueError):
        offset = 0

    if limit <= 0:
        limit = 200
    if limit > 500:
        limit = 500
    if offset < 0:
        offset = 0

    try:
        items = data_manager.load_strike_today_report_history(user_id, task_id, limit=limit, offset=offset)
        return jsonify({'success': True, 'task_id': task_id, 'items': items})
    except DatabaseError:
        logger.exception("Database error loading strike report history for task %s", task_id)
        return jsonify({'success': False, 'error': 'Database error loading strike report history'}), 503
    except Exception as e:
        logger.exception("Error loading strike report history for task %s", task_id)
        return jsonify({'success': False, 'error': str(e)}), 500


@task_bp.route('/<task_id>/complete', methods=['POST'])
@require_data_manager
@handle_database_error
def complete_task(task_id, user_id, data_manager):
    """Mark a task as completed for the authenticated user"""
    import time
    from src.services.performance_monitor import log_task_operation
    
    start_time = time.time()
    
    # Direct update instead of load-all/save-all
    success = data_manager.update_task_for_user(
        user_id,
        task_id,
        {
            'completed': True,
            'completed_at': datetime.now().isoformat(),
            'status': TaskStatus.COMPLETED.value
        }
    )
    
    if success:
        # Get updated task from database
        updated_task = data_manager.get_task_by_id(user_id, task_id)
        if updated_task:
            duration_ms = (time.time() - start_time) * 1000
            log_task_operation('complete', user_id, task_id, duration_ms, query_count=2)
            logger.info(f"Successfully completed task {task_id} for user {user_id}")
            return jsonify(updated_task)
        else:
            logger.error(f"Task {task_id} not found after update for user {user_id}")
            return jsonify({'error': 'Task not found after update'}), 500
    else:
        logger.error(f"Failed to complete task {task_id} for user {user_id}")
        return jsonify({'error': 'Failed to complete task'}), 500


@task_bp.route('/<task_id>/strike', methods=['POST'])
@require_data_manager
@require_json_body
@validate_input(validate_strike)
@handle_database_error
def strike_task(task_id, user_id, data_manager):
    """Unified strike endpoint for both today and forever.

    This endpoint also records aggregated strike analytics using the
    SQLite-backed analytics manager (no more JSON file writes).
    """
    import time
    from src.services.performance_monitor import log_task_operation
    
    strike_data = request.json
    strike_type = strike_data.get('type')
    report = strike_data.get('report', '')
    
    if report is None:
        report = ''
    
    start_time = time.time()
    
    # Get specific task instead of loading all
    task = data_manager.get_task_by_id(user_id, task_id)
    if not task:
        return jsonify({'error': 'Task not found'}), 404
    
    today = datetime.now().strftime('%Y-%m-%d')
    updates = {}
    
    if strike_type == 'today':
        # Check if task has already been struck twice today
        daily_strikes = task.get('daily_strikes', {})
        strikes_today = daily_strikes.get(today, 0)
        
        if strikes_today >= 2:
            return jsonify({'error': 'Maximum strikes reached for today'}), 400
        
        # Prepare updates for strike today
        strike_number = strikes_today + 1
        daily_strikes[today] = strike_number
        
        updates = {
            'daily_strikes': daily_strikes,
            'struck_today': True,
            'struck_date': today,
            'strike_report': report,
            'strike_count': task.get('strike_count', 0) + 1
        }
        
        # Compute recurrence snooze: hide task until its next occurrence.
        try:
            recurrence_type = (task.get('recurrence_type') or '').strip().lower()
            recurrence_param = task.get('recurrence_param')
            next_date = None
            if recurrence_type == 'every_n_days':
                try:
                    n = int(recurrence_param or 0)
                except Exception:  # noqa: broad-except
                    n = 0
                if n and n > 1:
                    base_dt = datetime.strptime(today, '%Y-%m-%d')
                    next_date = base_dt + timedelta(days=n)
            elif recurrence_type == 'weekly':
                try:
                    target_wd = int(recurrence_param)
                except Exception:  # noqa: broad-except
                    target_wd = None
                if target_wd is not None and 0 <= target_wd <= 6:
                    base_dt = datetime.strptime(today, '%Y-%m-%d')
                    days_ahead = (target_wd - base_dt.weekday()) % 7
                    if days_ahead == 0:
                        days_ahead = 7
                    next_date = base_dt + timedelta(days=days_ahead)
            # 'daily' / empty: clears naturally after daily reset; no snooze needed.
            if next_date is not None:
                updates['snoozed_until'] = next_date.strftime('%Y-%m-%d')
        except Exception:  # noqa: broad-except
            logger.exception("Failed to compute recurrence snooze for task %s at strike time", task_id)
        
        # Record event
        try:
            data_manager.add_strike_today_report_event(
                user_id=user_id,
                task_id=task_id,
                day=today,
                strike_number=strike_number,
                report=report,
            )
        except Exception:  # noqa: broad-except
            logger.exception("Failed to add strike_today report event")
    
    elif strike_type == 'forever':
        updates = {
            'completed': True,
            'completed_at': datetime.now().isoformat(),
            'struck_forever': True,
            'struck_today': True,
            'struck_date': today,
            'strike_report': report,
            'strike_count': task.get('strike_count', 0) + 1,
            'status': TaskStatus.COMPLETED.value
        }
    
    # Record strike event
    try:
        data_manager.add_strike_event(
            user_id=user_id,
            task_id=task_id,
            day=today,
            strike_type=strike_type,
        )
    except Exception:  # noqa: broad-except
        logger.exception("Failed to add strike event")
    
    # Direct update instead of load-all/save-all
    success = data_manager.update_task_for_user(user_id, task_id, updates)
    
    if success:
        updated_task = data_manager.get_task_by_id(user_id, task_id)
        if updated_task:
            duration_ms = (time.time() - start_time) * 1000
            log_task_operation('strike', user_id, task_id, duration_ms, query_count=2)
            # Increment analytics counters (decoupled from daily reset)
            try:
                from src.analytics_manager import increment_strike_counter
                increment_strike_counter()
            except Exception:  # noqa: broad-except
                # Non-fatal; analytics are best-effort
                logger.exception("Failed to increment strike counter")
            return jsonify(updated_task)
    
    return jsonify({'error': 'Failed to strike task'}), 500


@task_bp.route('/<task_id>/undo-strike', methods=['POST'])
@require_data_manager
@handle_database_error
def undo_strike(task_id, user_id, data_manager):
    """Undo a strike for today for the authenticated user"""
    import time
    from src.services.performance_monitor import log_task_operation

    start_time = time.time()
    
    # Get specific task instead of loading all
    task = data_manager.get_task_by_id(user_id, task_id)
    if not task:
        return jsonify({'error': 'Task not found'}), 404
    
    if not task.get('struck_today'):
        return jsonify({'error': 'Task is not struck for today'}), 400
    
    today = datetime.now().strftime('%Y-%m-%d')
    was_completed = task.get('completed', False)
    updates = {}
    
    if was_completed:
        # Undo strike forever - revert to incomplete state
        updates = {
            'completed': False,
            'completed_at': None,
            'struck_forever': False,
            'struck_today': False,
            'struck_date': None,
            'strike_report': None,
            'strike_count': max(0, task.get('strike_count', 0) - 1),
            'status': TaskStatus.PENDING.value
        }
    else:
        # Undo regular strike today
        daily_strikes = task.get('daily_strikes', {})
        strikes_today = daily_strikes.get(today, 0)
        
        if strikes_today > 0:
            daily_strikes[today] = strikes_today - 1
        
        updates = {
            'daily_strikes': daily_strikes,
            'struck_today': False if daily_strikes.get(today, 0) == 0 else True,
            'struck_date': None if daily_strikes.get(today, 0) == 0 else today,
            'strike_report': None if daily_strikes.get(today, 0) == 0 else task.get('strike_report'),
            'strike_count': max(0, task.get('strike_count', 0) - 1)
        }
    
    # Direct update instead of load-all/save-all
    success = data_manager.update_task_for_user(user_id, task_id, updates)
    
    if success:
        updated_task = data_manager.get_task_by_id(user_id, task_id)
        if updated_task:
            duration_ms = (time.time() - start_time) * 1000
            log_task_operation('undo_strike', user_id, task_id, duration_ms, query_count=2)
            logger.info(f"Successfully undid strike for task {task_id} for user {user_id}")
            return jsonify(updated_task)
        else:
            logger.error(f"Task {task_id} not found after undo-strike for user {user_id}")
            return jsonify({'error': 'Task not found after update'}), 500
    else:
        logger.error(f"Failed to undo strike for task {task_id} for user {user_id}")
        return jsonify({'error': 'Failed to undo strike'}), 500


@task_bp.route('/<task_id>/unschedule', methods=['POST'])
def unschedule_task(task_id):
    """Remove a task from the daily planner for the authenticated user"""
    user_id = _get_user_id()
    
    # Validate task_id format
    if not task_id or not isinstance(task_id, str) or len(task_id.strip()) == 0:
        return jsonify({'error': 'Invalid task ID'}), 400
    
    try:
        data_manager = _get_data_manager()
        if not data_manager:
            return jsonify({'error': 'Data manager not available'}), 500

        try:
            tasks = data_manager.load_tasks(user_id)
        except DatabaseError:
            logger.exception("Database error loading tasks for unschedule_task (user %s)", user_id)
            return jsonify({'error': 'Database error loading tasks'}), 503
        
        for i, task in enumerate(tasks):
            if task['id'] == task_id:
                # Remove scheduling information
                tasks[i]['scheduled_hour'] = None
                tasks[i]['scheduled_minute'] = None
                tasks[i]['scheduled_date'] = None
                tasks[i]['scheduled_duration'] = None
                
                if data_manager.save_tasks_for_user(user_id, tasks):
                    return jsonify(tasks[i])
                else:
                    return jsonify({'error': 'Failed to save tasks'}), 500
        
        return jsonify({'error': 'Task not found'}), 404
    except Exception:  # noqa: broad-except
        logger.exception("Error unscheduling task %s for user %s", task_id, user_id)
        return jsonify({'error': 'Internal server error'}), 500


@task_bp.route('/<task_id>/schedule', methods=['POST'])
@require_data_manager
@validate_input(validate_schedule)
@handle_database_error
def schedule_task(task_id, user_id, data_manager):
    """Schedule a task for a specific hour and duration for the authenticated user"""
    schedule_data = request.json
    if schedule_data is None:
        schedule_data = {}

    hour = schedule_data.get('hour')
    minute = schedule_data.get('minute', 0)
    duration = schedule_data.get('duration', 30)
    date = schedule_data.get('date', datetime.now().strftime('%Y-%m-%d'))

    tasks = data_manager.load_tasks(user_id)
    
    # Check for conflicts with existing scheduled tasks
    start_minutes = hour * 60 + minute
    end_minutes = start_minutes + duration
    
    for task in tasks:
        # Skip the task being scheduled and unscheduled tasks
        if task['id'] == task_id or not task.get('scheduled_date'):
            continue
        
        # Only check tasks on the same date
        if task.get('scheduled_date') != date:
            continue
        
        task_start_minutes = task.get('scheduled_hour', 0) * 60 + task.get('scheduled_minute', 0)
        task_duration = task.get('scheduled_duration', 30)
        task_end_minutes = task_start_minutes + task_duration
        
        # Check if there's an overlap
        if (start_minutes < task_end_minutes and end_minutes > task_start_minutes):
            conflict_hour = task.get('scheduled_hour')
            conflict_minute = task.get('scheduled_minute', 0)
            conflict_time = f"{conflict_hour}:{conflict_minute:02d}"
            
            logger.warning(f"Schedule conflict: Task {task_id} conflicts with {task['id']} at {conflict_time}")
            return jsonify({
                'error': 'Time slot conflict',
                'message': f"Conflicts with '{task['title']}' at {conflict_time}",
                'conflict_task': task['title'],
                'conflict_time': conflict_time
            }), 409
    
    for i, task in enumerate(tasks):
        if task['id'] == task_id:
            was_unscheduled = not task.get('scheduled_date')
            tasks[i]['scheduled_hour'] = hour
            tasks[i]['scheduled_minute'] = minute
            tasks[i]['scheduled_date'] = date
            tasks[i]['scheduled_duration'] = duration
            
            logger.info(f"Task {task_id} scheduled: hour={hour}, minute={minute}, date={date}, duration={duration}")
            
            if data_manager.save_tasks_for_user(user_id, tasks):
                if was_unscheduled:
                    try:
                        from src.analytics_manager import increment_analytics_counter
                        increment_analytics_counter('tasks_planned')
                    except Exception:  # noqa: broad-except
                        logger.exception("Failed to increment analytics counter (tasks_planned)")
                return jsonify(tasks[i])
            else:
                return jsonify({'error': 'Failed to save tasks'}), 500
    
    return jsonify({'error': 'Task not found'}), 404


@task_bp.route('/reset-daily-strikes', methods=['POST'])
def reset_daily_strikes():
    """Trigger a daily reset for the authenticated user via the scheduler service.

    Delegating to the centralized scheduler keeps behavior (refreshed_at,
    analytics, and reset logs) consistent whether the reset is automatic or
    user-initiated from the UI.
    """
    user_id = _get_user_id()
    data_manager = _get_data_manager()
    if not data_manager:
        return jsonify({'error': 'Data manager not available'}), 500

    try:
        try:
            # Ensure the scheduler service can resolve the data manager and
            # execute the canonical reset job for this user.
            scheduler_service.set_data_manager_getter(lambda: _get_data_manager())
            scheduler_service.reset_daily_strikes_job(replay=False, replay_reason='manual_api')
        except Exception:  # noqa: broad-except
            logger.exception("Scheduler daily reset job failed when triggered via API for user %s", user_id)
            return jsonify({'error': 'Failed to reset daily strikes'}), 500

        # For API callers we keep the simple success envelope that existing
        # frontend code expects. Detailed counts are available via analytics and
        # the /api/tasks/reset-log endpoint.
        return jsonify({'success': True, 'message': 'Daily reset completed'}), 200
    except Exception:  # noqa: broad-except
        logger.exception("Unexpected error in reset_daily_strikes for user %s", user_id)
        return jsonify({'error': 'Internal server error'}), 500


@task_bp.route('/export-excel', methods=['GET'])
def export_excel():
    """Export task report as Excel with custom date range.

    Query params:
    - start_date: YYYY-MM-DD (optional)
    - end_date: YYYY-MM-DD (optional)
    """
    user_id = _get_user_id()
    data_manager = _get_data_manager()
    if not data_manager:
        return jsonify({'error': 'Data manager not available'}), 500

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        tasks = data_manager.load_tasks(user_id)

        # Filter by created_at date range
        filtered = []
        for task in tasks:
            created = task.get('created_at', '')
            if created:
                try:
                    task_date = datetime.fromisoformat(created).strftime('%Y-%m-%d')
                except Exception:  # noqa: broad-except
                    task_date = ''
                if start_date and task_date < start_date:
                    continue
                if end_date and task_date > end_date:
                    continue
            filtered.append(task)

        # Load strike report history for each task
        task_reports = {}
        for task in filtered:
            try:
                reports = data_manager.load_strike_today_report_history(
                    user_id, task['id'], limit=50, offset=0
                )
                task_reports[task['id']] = reports
            except Exception:  # noqa: broad-except
                task_reports[task['id']] = []

        max_updates = max((len(r) for r in task_reports.values()), default=0)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Task Report'

        hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        hdr_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        cell_align = Alignment(vertical='center', wrap_text=True)
        done_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        headers = ['Date', 'Task Name', 'Category', 'Owner', 'Completed']
        for i in range(1, max_updates + 1):
            headers.append(f'Update {i}')

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = border

        for ridx, task in enumerate(filtered, 2):
            created = task.get('created_at', '')
            try:
                ds = datetime.fromisoformat(created).strftime('%d-%m-%Y') if created else ''
            except Exception:  # noqa: broad-except
                ds = created
            ws.cell(row=ridx, column=1, value=ds).border = border
            ws.cell(row=ridx, column=2, value=task.get('title', '')).border = border
            ws.cell(row=ridx, column=3, value=task.get('project', '')).border = border
            ws.cell(row=ridx, column=4, value=task.get('owner', '')).border = border

            is_done = bool(task.get('completed') or task.get('struck_forever'))
            ws.cell(row=ridx, column=5, value='✓' if is_done else '').border = border

            reports = task_reports.get(task['id'], [])
            for ui, rpt in enumerate(reports):
                text = f"{rpt.get('report', '')}"
                ws.cell(row=ridx, column=6 + ui, value=text).border = border

            for col in range(1, len(headers) + 1):
                c = ws.cell(row=ridx, column=col)
                c.alignment = cell_align
                if is_done:
                    c.fill = done_fill

        for letter, w in {'A': 14, 'B': 30, 'C': 18, 'D': 18, 'E': 12}.items():
            ws.column_dimensions[letter].width = w
        for i in range(max_updates):
            cl = chr(ord('F') + i) if i < 18 else None
            if cl:
                ws.column_dimensions[cl].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        dr = ''
        if start_date:
            dr += f'_from_{start_date}'
        if end_date:
            dr += f'_to_{end_date}'
        fname = f'Shakshuka_Report{dr}.xlsx'

        return send_file(
            buf, as_attachment=True, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception:  # noqa: broad-except
        logger.exception("Excel export failed for user %s", user_id)
        return jsonify({'error': 'Failed to export Excel report'}), 500


@task_bp.route('/reset-log', methods=['GET'])
def get_reset_log():
    """Return the latest unseen daily reset log for the current user, if any."""
    user_id = _get_user_id()
    data_manager = _get_data_manager()
    if not data_manager:
        return jsonify({'success': False, 'error': 'Data manager not available'}), 500

    try:
        log_entry = data_manager.get_latest_daily_reset_log(user_id, include_seen=False)
        if not log_entry:
            return jsonify({'success': True, 'log': None}), 200

        # The data manager already returns a safe, compact payload.
        return jsonify({'success': True, 'log': log_entry}), 200
    except DatabaseError:
        logger.exception("Database error loading daily reset log for user %s", user_id)
        return jsonify({'success': False, 'error': 'Database error loading reset log'}), 503
    except Exception as e:  # noqa: broad-except
        logger.exception("Unexpected error loading daily reset log for user %s", user_id)
        return jsonify({'success': False, 'error': str(e)}), 500


@task_bp.route('/reset-log/clear', methods=['POST'])
def clear_reset_log():
    """Mark the latest (or a specific) daily reset log as seen for the current user."""
    user_id = _get_user_id()
    data_manager = _get_data_manager()
    if not data_manager:
        return jsonify({'success': False, 'error': 'Data manager not available'}), 500

    payload = request.json if request.is_json else None
    log_id = None
    if isinstance(payload, dict):
        try:
            log_id_raw = payload.get('id')
            if log_id_raw is not None:
                log_id = int(log_id_raw)
        except (TypeError, ValueError):
            log_id = None

    try:
        ok = data_manager.mark_daily_reset_log_seen(user_id, log_id=log_id)
        if not ok:
            # Not fatal; simply indicate that nothing changed so callers can
            # clear local state.
            return jsonify({'success': True, 'updated': False}), 200
        return jsonify({'success': True, 'updated': True}), 200
    except DatabaseError:
        logger.exception("Database error marking daily reset log seen for user %s", user_id)
        return jsonify({'success': False, 'error': 'Database error updating reset log'}), 503
    except Exception as e:  # noqa: broad-except
        logger.exception("Unexpected error marking daily reset log seen for user %s", user_id)
        return jsonify({'success': False, 'error': str(e)}), 500
